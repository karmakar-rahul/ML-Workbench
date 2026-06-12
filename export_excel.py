"""
export_excel.py

Excel export module for ML Workbench Dashboard.

Creates a detailed Excel report containing:

1. Dataset Summary
2. Raw Data
3. Train Data
4. Test Data
5. Predictions
6. Metrics
7. Classification Report
8. Confusion Matrix
9. Feature Importance
10. Cross Validation
11. Model Parameters


"""

import pandas as pd

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)

from openpyxl.utils import get_column_letter


# STYLING

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True
)

CENTER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center"
)


# FORMAT WORKSHEET

def format_worksheet(ws):
    """
    Apply formatting to worksheet.
    """

    for cell in ws[1]:

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT

    ws.freeze_panes = "A2"

    for column in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column[0].column
        )

        for cell in column:

            try:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
            except Exception:
                pass

        adjusted_width = min(
            max_length + 3,
            50
        )

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width


# WRITE DATAFRAME

def write_dataframe(
    writer,
    dataframe,
    sheet_name
):
    """
    Safely write dataframe.
    """

    if dataframe is None:
        return

    if len(dataframe) == 0:
        return

    dataframe.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False
    )


# MODEL PARAMETERS

def model_parameters_dataframe(
    model
):
    """
    Convert model parameters to dataframe.
    """

    try:

        params = model.get_params()

        return pd.DataFrame({
            "Parameter":
                list(params.keys()),
            "Value":
                list(params.values())
        })

    except Exception:

        return pd.DataFrame()


# CROSS VALIDATION DF

def cross_validation_dataframe(
    cv_scores
):
    """
    CV dataframe.
    """

    if cv_scores is None:
        return pd.DataFrame()

    return pd.DataFrame({

        "Fold":
            list(
                range(
                    1,
                    len(cv_scores) + 1
                )
            ),

        "Score":
            cv_scores
    })


# DATASET SUMMARY DF

def dataset_summary_dataframe(
    summary_dict
):
    """
    Summary dictionary -> dataframe.
    """

    if summary_dict is None:
        return pd.DataFrame()

    return pd.DataFrame({

        "Metric":
            list(summary_dict.keys()),

        "Value":
            list(summary_dict.values())
    })


# MAIN EXPORT FUNCTION

def export_ml_report(
    output_path,

    raw_data=None,
    train_data=None,
    test_data=None,

    predictions=None,

    metrics_df=None,
    classification_report_df=None,
    confusion_matrix_df=None,

    feature_importance_df=None,
    cross_validation_scores=None,

    dataset_summary=None,
    model=None
):
    """
    Generate complete Excel report.
    """

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:

        
        # Dataset Summary
        

        summary_df = (
            dataset_summary_dataframe(
                dataset_summary
            )
        )

        write_dataframe(
            writer,
            summary_df,
            "Dataset_Summary"
        )

        
        # Raw Data
        

        write_dataframe(
            writer,
            raw_data,
            "Raw_Data"
        )

        
        # Train Data
        

        write_dataframe(
            writer,
            train_data,
            "Train_Data"
        )

        
        # Test Data
        

        write_dataframe(
            writer,
            test_data,
            "Test_Data"
        )

        
        # Predictions
        

        write_dataframe(
            writer,
            predictions,
            "Predictions"
        )

        
        # Metrics
        

        write_dataframe(
            writer,
            metrics_df,
            "Metrics"
        )

        
        # Classification Report
        

        write_dataframe(
            writer,
            classification_report_df,
            "Classification_Report"
        )

        
        # Confusion Matrix
        

        write_dataframe(
            writer,
            confusion_matrix_df,
            "Confusion_Matrix"
        )

        
        # Feature Importance
        

        write_dataframe(
            writer,
            feature_importance_df,
            "Feature_Importance"
        )

        
        # Cross Validation
        

        cv_df = cross_validation_dataframe(
            cross_validation_scores
        )

        write_dataframe(
            writer,
            cv_df,
            "Cross_Validation"
        )

        
        # Model Parameters
        

        params_df = (
            model_parameters_dataframe(
                model
            )
        )

        write_dataframe(
            writer,
            params_df,
            "Model_Parameters"
        )

        
        # Format Sheets
        

        workbook = writer.book

        for sheet_name in workbook.sheetnames:

            ws = workbook[sheet_name]

            format_worksheet(ws)

    return output_path


# QUICK EXPORT WRAPPER

def export_classification_report(
    output_path,
    raw_data,
    train_data,
    test_data,
    predictions,
    metrics_df,
    classification_report_df,
    confusion_matrix_df,
    feature_importance_df,
    dataset_summary,
    model,
    cv_scores=None
):
    """
    Classification shortcut.
    """

    return export_ml_report(
        output_path=output_path,

        raw_data=raw_data,
        train_data=train_data,
        test_data=test_data,

        predictions=predictions,

        metrics_df=metrics_df,

        classification_report_df=
        classification_report_df,

        confusion_matrix_df=
        confusion_matrix_df,

        feature_importance_df=
        feature_importance_df,

        cross_validation_scores=
        cv_scores,

        dataset_summary=
        dataset_summary,

        model=model
    )


# REGRESSION SHORTCUT

def export_regression_report(
    output_path,
    raw_data,
    train_data,
    test_data,
    predictions,
    metrics_df,
    feature_importance_df,
    dataset_summary,
    model,
    cv_scores=None
):
    """
    Regression shortcut.
    """

    return export_ml_report(
        output_path=output_path,

        raw_data=raw_data,
        train_data=train_data,
        test_data=test_data,

        predictions=predictions,

        metrics_df=metrics_df,

        feature_importance_df=
        feature_importance_df,

        cross_validation_scores=
        cv_scores,

        dataset_summary=
        dataset_summary,

        model=model
    )


# CLUSTERING SHORTCUT

def export_clustering_report(
    output_path,
    raw_data,
    metrics_df,
    dataset_summary,
    model
):
    """
    Clustering shortcut.
    """

    return export_ml_report(
        output_path=output_path,

        raw_data=raw_data,

        metrics_df=metrics_df,

        dataset_summary=
        dataset_summary,

        model=model
    )